"""
transform_badges.py
====================
Reads all CSV files in INPUT_FOLDER, transforms badge/training data into
the master tracker format, appends new (non-duplicate) rows to the
"Masterlist" sheet of MASTER_FILE, and writes a log entry for every run.
 
All other sheets in the master file are left completely untouched.
Existing rows, formatting, and styles in "Masterlist" are preserved —
new rows are appended after the last occupied row only.
 
After successful processing, each CSV is moved to PROCESSED_FOLDER.
 
USAGE
-----
1. Set the four paths in the CONFIG section below.
2. Run:  python transform_badges.py
 
ASSUMPTIONS
-----------
- Input CSVs follow the column structure documented in SOURCE COLUMNS below.
  Column *names* may vary (see COLUMN_ALIASES); the script matches by common
  aliases, not exact names.
- The master file already exists and contains a sheet named exactly "Masterlist"
  with a header row. If the file does not exist, a new one is created with
  just that sheet and its header.
- A row is considered a duplicate if it shares the same Email + Skills Area
  and Level + Date of Award as an existing master row.
- Date values may arrive as  "22-Aug-25", "2025-08-22", a plain number
  (Excel serial), or other common formats — all are normalised to "DD-MMM-YY".
"""

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG  ← edit these four paths before running
# ─────────────────────────────────────────────────────────────────────────────

INPUT_FOLDER     = r"C:\Users\incoming"           # folder containing input CSVs
PROCESSED_FOLDER = r"C:\Users\processed"       # CSVs are moved here after processing
MASTER_FILE      = r"C:\Users\mock_master_list.xlsx"  # master Excel tracker file
LOG_FILE         = r"C:\Users\transform_log.xlsx" # append-only log

# Name of the sheet inside MASTER_FILE where rows are appended
MASTERLIST_SHEET = "Masterlist"

# Row number (1-based) of the header row in the Masterlist sheet
# Change this if your headers are not on row 1 (e.g. set to 3 if headers are on row 3)
HEADER_ROW = 3

# ─────────────────────────────────────────────────────────────────────────────
# IMPORTS
# ─────────────────────────────────────────────────────────────────────────────
 
import os
import re
import glob
import shutil
import logging
import traceback
from datetime import datetime, timedelta
 
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment as XlAlign
 
# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
 
# Master file output columns (order matters — matches the tracker header row)
MASTER_COLUMNS = [
    "Name", "Email", "Training Provider",
    "Skills Area", "Skills Area and Level", "Badge Level",
    "Date of Award", "Month", "Year", "Programme",
    # The columns below exist in the master file but are NOT populated by this
    # script (manual / CRM fields).  They are preserved when appending.
    "Upload CRM", "Column13", "DTO Processed Date (Sep 25 start)",
    "Remarks", "Column17", "S/N. (Dec 25 start)",
]
 
# Columns this script actively writes (subset of MASTER_COLUMNS)
WRITE_COLUMNS = [
    "Name", "Email", "Training Provider",
    "Skills Area", "Skills Area and Level", "Badge Level",
    "Date of Award", "Month", "Year", "Programme",
]
 
# Duplicate-detection key: a row is a duplicate if all three match
DEDUP_KEY = ["Email", "Skills Area and Level", "Date of Award"]
 
# ── Flexible column name mapping ──────────────────────────────────────────────
# Keys are the *canonical* names used inside this script.
# Values are lists of alternative spellings found in real CSVs (case-insensitive).
COLUMN_ALIASES: dict[str, list[str]] = {
    "Email": [
        "email address", "email", "identifier (email_address)",
        "identifier", "emailaddress",
    ],
    "Preferred Name": [
        "preferred name(name to appear on badge)",
        "preferred name", "preferredname", "name on badge", "display name",
    ],
    "Skills Badge": [
        "name of skills badge", "skills badge", "badge name",
        "skillsbadge", "badge",
    ],
    "Programme": [
        "course / programme title", "programme title", "course title",
        "programme", "course", "coursetitle",
    ],
    "Date": [
        "date of course completion", "completion date", "date completed",
        "course date", "date of completion", "date",
    ],
    "Training Provider": [
        "training provider", "provider", "trainingprovider",
        "training organisation", "organisation",
    ],
}
 
# Month abbreviations for date formatting
MONTH_ABBR = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]
 
# ─────────────────────────────────────────────────────────────────────────────
# LOGGING SETUP  (console + log file)
# ─────────────────────────────────────────────────────────────────────────────
 
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)
 
 
def _append_log(file: str, timestamp: str, filename: str,
                rows_added: int, status: str, detail: str = "") -> None:
    """Append one row to the Excel log file.
 
    Uses openpyxl to read-then-append, so the file is a proper .xlsx with
    no encoding or line-ending issues.  Creates the file with a styled
    header row if it does not exist yet.
    """
    from openpyxl.styles import Font, PatternFill, Alignment as XlAlignLog
 
    LOG_HEADERS = ["Timestamp", "File Name", "Rows Added", "Status", "Detail"]
    new_row     = [timestamp, filename, rows_added, status, detail]
 
    if os.path.isfile(file):
        wb = load_workbook(file)
        ws = wb.active
    else:
        # Create fresh workbook with a formatted header row
        wb = Workbook()
        ws = wb.active
        ws.title = "Log"
        ws.append(LOG_HEADERS)
 
        # Bold + light-blue header
        header_fill = PatternFill("solid", fgColor="BDD7EE")
        for cell in ws[1]:
            cell.font      = Font(bold=True)
            cell.fill      = header_fill
            cell.alignment = XlAlignLog(horizontal="center")
 
        # Freeze header row so it stays visible while scrolling
        ws.freeze_panes = "A2"
 
        # Set sensible column widths
        for col_letter, width in zip("ABCDE", [22, 52, 12, 12, 60]):
            ws.column_dimensions[col_letter].width = width
 
    ws.append(new_row)
    wb.save(file)
 
 
# ─────────────────────────────────────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────
 
def _normalise_col_name(name: str) -> str:
    """Lower-case, strip, collapse internal whitespace."""
    return re.sub(r"\s+", " ", str(name).strip().lower())
 
 
def _map_columns(df: pd.DataFrame) -> dict[str, str]:
    """
    Return a dict  {canonical_name: actual_df_column}  for every canonical
    column we need.  Raises ValueError if a required column cannot be found.
    """
    normalised = {_normalise_col_name(c): c for c in df.columns}
    mapping: dict[str, str] = {}
 
    for canonical, aliases in COLUMN_ALIASES.items():
        found = None
        for alias in aliases:
            if alias in normalised:
                found = normalised[alias]
                break
        if found is None:
            # Last resort: partial match
            for norm_col, real_col in normalised.items():
                if any(alias.split()[0] in norm_col for alias in aliases):
                    found = real_col
                    break
        if found is None:
            raise ValueError(
                f"Cannot find column '{canonical}'. "
                f"Checked aliases: {aliases}. "
                f"Available columns: {list(df.columns)}"
            )
        mapping[canonical] = found
 
    return mapping
 
 
def _excel_serial_to_date(serial: float) -> datetime | None:
    """Convert an Excel date serial number (e.g. 45912) to a Python datetime."""
    try:
        epoch = datetime(1899, 12, 30)
        return epoch + timedelta(days=float(serial))
    except Exception:
        return None
 
 
def _parse_date(raw) -> datetime | None:
    """
    Try to parse *raw* (string, int, float, or datetime) into a Python datetime.
    Returns None if parsing fails.
    """
    if pd.isna(raw):
        return None
 
    # Already a datetime / Timestamp
    if isinstance(raw, (datetime, pd.Timestamp)):
        return pd.Timestamp(raw).to_pydatetime()
 
    # Numeric  → Excel serial
    if isinstance(raw, (int, float)):
        return _excel_serial_to_date(raw)
 
    raw_str = str(raw).strip()
    if not raw_str:
        return None
 
    # Try numeric string (Excel serial stored as text)
    try:
        return _excel_serial_to_date(float(raw_str))
    except ValueError:
        pass
 
    # Common explicit formats (add more here if new formats appear in the wild)
    for fmt in (
        "%d-%b-%y",   # 22-Aug-25
        "%d-%b-%Y",   # 22-Aug-2025
        "%d/%m/%Y",   # 22/08/2025
        "%d/%m/%y",   # 22/08/25
        "%Y-%m-%d",   # 2025-08-22
        "%m/%d/%Y",   # 08/22/2025
        "%d %b %Y",   # 22 Aug 2025
        "%d %B %Y",   # 22 August 2025
        "%B %d, %Y",  # August 22, 2025
    ):
        try:
            return datetime.strptime(raw_str, fmt)
        except ValueError:
            pass
 
    # pandas last-ditch parse
    try:
        return pd.to_datetime(raw_str, dayfirst=True).to_pydatetime()
    except Exception:
        return None
 
 
def _format_date(raw) -> tuple[str, str, str]:
    """
    Return (display_string, month_number, year_number) for *raw* date value.
    display_string → "DD-MMM-YY"  e.g. "22-Aug-25"
    month_number   → "8"
    year_number    → "2025"
    If parsing fails, returns ("", "", "").
    """
    dt = _parse_date(raw)
    if dt is None:
        return "", "", ""
    dd  = str(dt.day).zfill(2)
    mmm = MONTH_ABBR[dt.month - 1]
    yy  = str(dt.year)[-2:]
    return f"{dd}-{mmm}-{yy}", str(dt.month), str(dt.year)
 
 
def _parse_badge(badge_str: str) -> tuple[str, str, str]:
    """
    Split  "HR Consulting (Basic)"  into:
        skills_area         → "HR Consulting"
        skills_area_level   → "HR Consulting (Basic)"
        badge_level         → "Basic"
 
    If no parentheses are found, badge_level is set to the value in the
    "Skills Badge Level" column (if available) or left blank.
    """
    badge_str = str(badge_str).strip()
    match = re.match(r"^(.*?)\s*\(([^)]+)\)\s*$", badge_str)
    if match:
        return match.group(1).strip(), badge_str, match.group(2).strip()
    return badge_str, badge_str, ""
 
 
# ─────────────────────────────────────────────────────────────────────────────
# CORE TRANSFORMATION
# ─────────────────────────────────────────────────────────────────────────────
 
def transform_csv(filepath: str) -> pd.DataFrame:
    """
    Read one CSV file and return a DataFrame with WRITE_COLUMNS columns.
 
    Transformation steps
    --------------------
    1. Read CSV (try UTF-8-sig first, then latin-1 as fallback).
    2. Map source columns to canonical names using COLUMN_ALIASES.
    3. Parse and reformat the date field.
    4. Split the Skills Badge name into Area / Area+Level / Level.
    5. Normalise Email to upper-case (matches master file convention).
    6. Drop rows that are completely blank (no email, no name, no badge).
    7. Return only the WRITE_COLUMNS subset.
    """
    # ── Read ────────────────────────────────────────────────────────────────
    try:
        df = pd.read_csv(filepath, encoding="utf-8-sig", dtype=str)
    except UnicodeDecodeError:
        df = pd.read_csv(filepath, encoding="latin-1", dtype=str)
 
    df.columns = [c.strip() for c in df.columns]  # strip whitespace from headers
 
    # ── Map columns ─────────────────────────────────────────────────────────
    col = _map_columns(df)  # {canonical: real_column_name}
 
    # ── Build output rows ────────────────────────────────────────────────────
    records = []
    for _, row in df.iterrows():
        email     = str(row[col["Email"]]).strip().upper()
        name      = str(row[col["Preferred Name"]]).strip()
        badge_raw = str(row[col["Skills Badge"]]).strip()
        programme = str(row[col["Programme"]]).strip()
        date_raw  = row[col["Date"]]
        provider  = str(row[col["Training Provider"]]).strip()
 
        # Drop entirely blank rows
        if not email or not name or not badge_raw or email == "NAN":
            continue
 
        # Badge parsing — if badge text has no parentheses, fall back to
        # the "Skills Badge Level" column (column F in the sample CSV)
        skills_area, skills_area_level, badge_level = _parse_badge(badge_raw)
        if not badge_level:
            # Try to find a badge level column in the source
            for col_name in df.columns:
                if "level" in col_name.lower() and "badge" in col_name.lower():
                    fallback = str(row[col_name]).strip()
                    if fallback and fallback.upper() != "NAN":
                        badge_level = fallback
                        # Rebuild skills_area_level with the found level
                        skills_area_level = f"{skills_area} ({badge_level})"
                    break
 
        date_display, month, year = _format_date(date_raw)
 
        records.append({
            "Name":                  name,
            "Email":                 email,
            "Training Provider":     provider,
            "Skills Area":           skills_area,
            "Skills Area and Level": skills_area_level,
            "Badge Level":           badge_level,
            "Date of Award":         date_display,
            "Month":                 month,
            "Year":                  year,
            "Programme":             programme,
        })
 
    return pd.DataFrame(records, columns=WRITE_COLUMNS)
 
 
# ─────────────────────────────────────────────────────────────────────────────
# MASTER FILE HELPERS
# ─────────────────────────────────────────────────────────────────────────────
 
def _read_masterlist_for_dedup(path: str) -> tuple[dict[str, int], set[tuple]]:
    """
    Open the master file with openpyxl and read ONLY the Masterlist sheet.
 
    Returns
    -------
    col_index : dict {header_name: 0-based column index}
        Maps each header string to its column position in the sheet.
    existing_keys : set of tuples
        The dedup keys already present so we can skip duplicates.
 
    If the file does not exist, returns empty structures — the caller will
    create the file from scratch when it tries to append.
    """
    if not os.path.isfile(path):
        logger.warning("Master file not found — it will be created: %s", path)
        return {}, set()
 
    wb = load_workbook(path, read_only=True, data_only=True)
 
    if MASTERLIST_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f'Sheet "{MASTERLIST_SHEET}" not found in {path}. '
            f"Available sheets: {wb.sheetnames}"
        )
 
    ws = wb[MASTERLIST_SHEET]
 
    # Read the configured header row to build column index
    header_row = next(
        ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True),
        None,
    )
    if not header_row:
        wb.close()
        return {}, set()
 
    col_index = {str(v).strip(): i for i, v in enumerate(header_row) if v is not None}
 
    # Identify which positions hold the dedup key columns
    key_positions = []
    for k in DEDUP_KEY:
        pos = col_index.get(k)
        if pos is None:
            logger.warning("Dedup column '%s' not found in master header — skipping.", k)
        key_positions.append(pos)
 
    # Collect existing dedup keys from all data rows (everything after the header)
    existing_keys: set[tuple] = set()
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        key = tuple(
            str(row[p]).strip().lower() if (p is not None and p < len(row) and row[p] is not None) else ""
            for p in key_positions
        )
        existing_keys.add(key)
 
    wb.close()
    return col_index, existing_keys
 
 
def _append_rows_to_master(path: str, new_rows: pd.DataFrame) -> None:
    """
    Append *new_rows* to the Masterlist sheet of *path* using openpyxl.
 
    Rules
    -----
    - Only the Masterlist sheet is touched; all other sheets are unchanged.
    - Existing rows, formulas, and formatting in Masterlist are preserved.
    - New rows are written after the last non-empty row.
    - Column order follows the header already in the sheet (not MASTER_COLUMNS),
      so the data always lands in the right columns regardless of header order.
    - Date of Award, Month, Year cells are centre-aligned.
    - If the master file does not exist, it is created with a fresh Masterlist
      sheet using WRITE_COLUMNS as the header.
    """
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
 
    # ── Load (or create) the workbook ────────────────────────────────────────
    if os.path.isfile(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
        # Remove default "Sheet" and create Masterlist
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        ws = wb.create_sheet(MASTERLIST_SHEET)
        # Write header from WRITE_COLUMNS
        for col_idx, header in enumerate(WRITE_COLUMNS, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        wb.save(path)
        wb = load_workbook(path)
 
    if MASTERLIST_SHEET not in wb.sheetnames:
        raise ValueError(
            f'Sheet "{MASTERLIST_SHEET}" not found in {path}. '
            f"Available sheets: {wb.sheetnames}"
        )
 
    ws = wb[MASTERLIST_SHEET]
 
    # ── Read the sheet header to know which column each field goes into ───────
    header_row = [cell.value for cell in ws[HEADER_ROW]]
    # Build {field_name: 1-based Excel column number}
    sheet_col = {
        str(v).strip(): idx + 1
        for idx, v in enumerate(header_row)
        if v is not None
    }
 
    # Columns that should be centre-aligned
    centre_field_names = {"Date of Award", "Month", "Year"}
    centre_cols = {sheet_col[f] for f in centre_field_names if f in sheet_col}
 
    # ── Find the first truly empty row after the data ─────────────────────────
    # Anchor on the "Name" column's actual Excel column number — NOT hardcoded
    # column 1, because your data may not start in column A.
    # Fallback chain: Name → Email → first header column found.
    anchor_col = sheet_col.get("Name") or sheet_col.get("Email") or min(sheet_col.values())
 
    last_data_row = HEADER_ROW  # fallback: no data yet, write right after header
    for r in range(ws.max_row, HEADER_ROW, -1):
        if ws.cell(row=r, column=anchor_col).value is not None:
            last_data_row = r
            break
    next_row = last_data_row + 1
 
    logger.debug("  anchor_col=%d  last_data_row=%d  next_row=%d",
                 anchor_col, last_data_row, next_row)
 
    # ── Write each new row ────────────────────────────────────────────────────
    for _, data_row in new_rows.iterrows():
        for field, value in data_row.items():
            col_num = sheet_col.get(str(field).strip())
            if col_num is None:
                continue  # field not in sheet header — skip silently
            cell = ws.cell(row=next_row, column=col_num, value=value if value != "" else None)
            if col_num in centre_cols:
                cell.alignment = XlAlign(horizontal="center")
        next_row += 1
 
    wb.save(path)
 
 
# ─────────────────────────────────────────────────────────────────────────────
# MAIN PIPELINE
# ─────────────────────────────────────────────────────────────────────────────
 
def process_all() -> None:
    """
    Find every CSV in INPUT_FOLDER, transform it, deduplicate against the
    Masterlist sheet, append only new rows, move the CSV to PROCESSED_FOLDER,
    and log the result.
    """
    csv_files = glob.glob(os.path.join(INPUT_FOLDER, "*.csv"))
    if not csv_files:
        logger.info("No CSV files found in %s — nothing to do.", INPUT_FOLDER)
        return
 
    # Create the processed folder if it doesn't exist yet
    os.makedirs(PROCESSED_FOLDER, exist_ok=True)
 
    # ── Read existing master data once (for dedup) ────────────────────────────
    # We use read_only mode so the file is not locked during the loop.
    col_index, existing_keys = _read_masterlist_for_dedup(MASTER_FILE)
 
    # Helper: build a dedup key tuple from a DataFrame row.
    # Use row[k] with a try/except — pandas Series supports .get() but it can
    # silently return None for missing labels; this is explicit and safe.
    def _key(row) -> tuple:
        def _val(k):
            try:
                v = row[k]
                return str(v).strip().lower() if v is not None and str(v).upper() != "NAN" else ""
            except KeyError:
                return ""
        return tuple(_val(k) for k in DEDUP_KEY)
 
    total_added = 0
 
    for filepath in sorted(csv_files):
        filename  = os.path.basename(filepath)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        logger.info("Processing: %s", filename)
        moved = False
 
        try:
            transformed = transform_csv(filepath)
 
            if transformed.empty:
                logger.warning("  No valid rows found in %s", filename)
                _append_log(LOG_FILE, timestamp, filename, 0, "SKIPPED", "No valid rows")
                # Still move the file — it was read successfully, just had no data
                _move_to_processed(filepath, PROCESSED_FOLDER)
                moved = True
                continue
 
            # ── Deduplicate ───────────────────────────────────────────────────
            new_rows = transformed[
                transformed.apply(_key, axis=1).apply(lambda k: k not in existing_keys)
            ].copy()
 
            duplicate_count = len(transformed) - len(new_rows)
            if duplicate_count:
                logger.info("  Skipped %d duplicate row(s) already in master.", duplicate_count)
                # Log which rows were considered duplicates to help diagnose false positives
                dupes = transformed[
                    transformed.apply(_key, axis=1).apply(lambda k: k in existing_keys)
                ]
                for _, dr in dupes.iterrows():
                    logger.debug("    DUPE: %s | %s | %s",
                                 dr.get("Email","?"), dr.get("Skills Area and Level","?"),
                                 dr.get("Date of Award","?"))
 
            if new_rows.empty:
                logger.info("  All rows already exist in master — nothing appended.")
                _append_log(LOG_FILE, timestamp, filename, 0,
                            "SKIPPED", "All rows already in master")
                _move_to_processed(filepath, PROCESSED_FOLDER)
                moved = True
                continue
 
            # Keep only the columns this script writes; leave manual columns blank
            new_rows = new_rows[WRITE_COLUMNS]
 
            # ── Append to the Masterlist sheet (safe — no other sheets touched) ──
            _append_rows_to_master(MASTER_FILE, new_rows)
 
            # Update the in-memory dedup set so subsequent files in this run
            # are also checked against the rows we just added.
            for _, r in new_rows.iterrows():
                existing_keys.add(_key(r))
 
            rows_added = len(new_rows)
            total_added += rows_added
            logger.info("  Appended %d new row(s).", rows_added)
            _append_log(
                LOG_FILE, timestamp, filename, rows_added, "SUCCESS",
                f"{duplicate_count} duplicate(s) skipped",
            )
 
            # ── Move the processed CSV ────────────────────────────────────────
            _move_to_processed(filepath, PROCESSED_FOLDER)
            moved = True
 
        except Exception as exc:
            detail = f"{type(exc).__name__}: {exc}"
            logger.error("  FAILED — %s", detail)
            logger.debug(traceback.format_exc())
            _append_log(LOG_FILE, timestamp, filename, 0, "ERROR", detail)
            # Do NOT move the file on error so the user can inspect and retry
 
    if total_added > 0:
        logger.info(
            "Done. %d new row(s) total appended to sheet '%s' in %s",
            total_added, MASTERLIST_SHEET, MASTER_FILE,
        )
    else:
        logger.info("No new rows to write — master file unchanged.")
 
 
def _move_to_processed(src: str, dest_folder: str) -> None:
    """
    Move *src* file into *dest_folder*.
    If a file with the same name already exists there, a timestamp suffix is
    added to avoid silently overwriting it.
    """
    filename = os.path.basename(src)
    dest = os.path.join(dest_folder, filename)
 
    # Avoid overwriting an existing file in the processed folder
    if os.path.exists(dest):
        stem, ext = os.path.splitext(filename)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = os.path.join(dest_folder, f"{stem}_{stamp}{ext}")
 
    shutil.move(src, dest)
    logger.info("  Moved to processed: %s", os.path.basename(dest))
 
 
# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
 
if __name__ == "__main__":
    process_all()
