"""
badge_pipeline.py
==================
Full badge processing pipeline — runs in this order:

  PART 1 — transform_badges   : Read CSVs from incoming folder, transform and
                                 append new rows to the Masterlist sheet.
  PART 2 — award_role_badges  : Check Masterlist for users who have earned all
                                 required skill badges for a role badge and
                                 append any missing role badge rows.
  PART 3 — export outputs     : Copy the Excel badge template to a timestamped
                                 To_DTO folder, fill it with all new rows from
                                 Parts 1 & 2, and also export those rows as a
                                 combined CSV in the same folder.

A single Excel log file (pipeline_log.xlsx) records every action.

USAGE
-----
1. Edit the CONFIG section below.
2. Run:  python badge_pipeline.py
"""

# =============================================================================
# CONFIG  ← edit all paths and settings here before running
# =============================================================================

# ── Folders ──────────────────────────────────────────────────────────────────
INPUT_FOLDER        = r"C:\Users\incoming"
PROCESSED_FOLDER    = r"C:\Users\processed"
TO_DTO_BASE_FOLDER  = r"C:\Users\To_DTO"

# ── Files ─────────────────────────────────────────────────────────────────────
MASTER_FILE         = r"C:\Users\mock_master_list.xlsx"
LOG_FILE            = r"C:\Users\pipeline_log.xlsx"
BADGE_TEMPLATE_FILE = r"C:\Users\Badges_Excel_Template.xlsx"

# ── Master sheet settings ─────────────────────────────────────────────────────
MASTERLIST_SHEET    = "Masterlist"
HEADER_ROW          = 3       # 1-based row number of header in Masterlist sheet

# ── Role badge settings ───────────────────────────────────────────────────────
ROLE_BADGE_PROVIDER = "company"

# ── Processing mode ───────────────────────────────────────────────────────────
PROCESS_MODE = "both"   # "skills" | "roles" | "both"

# ── Export split mode ─────────────────────────────────────────────────────────
# EXPORT_MODE controls how rows are grouped into output files:
EXPORT_MODE = "combined"   # "combined" | "split"

# ── Role badge → required skills mapping ─────────────────────────────────────
# Keys   : role badge name used as the dict key (for internal logic)
# Values : list of required "Skills Area" values (ALL must be present)
ROLE_BADGE_REQUIREMENTS: dict[str, list[str]] = {
    "Example Role Badge": [
        "Skills Badge A",
        "Skills Badge B",
    ],
}

# =============================================================================
# IMPORTS
# =============================================================================

import os
import re
import glob
import shutil
import logging
import traceback
from copy import copy
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# =============================================================================
# LOGGING SETUP
# =============================================================================

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# =============================================================================
# CONSTANTS
# =============================================================================

WRITE_COLUMNS = [
    "Name", "Email", "Training Provider",
    "Skills Area", "Skills Area and Level", "Badge Level",
    "Date of Award", "Month", "Year", "Programme",
]

DEDUP_KEY = ["Email", "Skills Area and Level", "Date of Award"]

COLUMN_ALIASES: dict[str, list[str]] = {
    "Email": [
        "email address", "email", "identifier (email_address)",
        "identifier", "emailaddress",
    ],
    "Preferred Name": [
        "preferred name(name to appear on badge)",
        "preferred name", "preferredname", "name on badge", "display name",
    ],
    "Skills Badge": [
        "name of skills badge", "skills badge", "badge name",
        "skillsbadge", "badge",
    ],
    "Programme": [
        "course / programme title", "programme title", "course title",
        "programme", "course", "coursetitle",
    ],
    "Date": [
        "date of course completion", "completion date", "date completed",
        "course date", "date of completion", "date",
    ],
    "Training Provider": [
        "training provider", "provider", "trainingprovider",
        "training organisation", "organisation",
    ],
}

MONTH_ABBR = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]

# =============================================================================
# SHARED UTILITIES
# =============================================================================

# ── Excel log ─────────────────────────────────────────────────────────────────

def _append_log(timestamp: str, part: str, subject: str,
                rows: int, status: str, detail: str = "") -> None:
    """Append one row to the pipeline Excel log file."""
    headers = ["Timestamp", "Part", "Subject", "Rows", "Status", "Detail"]
    exists  = os.path.isfile(LOG_FILE)

    if exists:
        wb = load_workbook(LOG_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "PipelineLog"
        ws.append(headers)
        fill = PatternFill("solid", fgColor="BDD7EE")
        for cell in ws[1]:
            cell.font      = Font(bold=True)
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        for col, width in zip("ABCDEF", [22, 10, 52, 8, 12, 80]):
            ws.column_dimensions[col].width = width

    ws.append([timestamp, part, subject, rows, status, detail])
    wb.save(LOG_FILE)


def _ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ── Date helpers ──────────────────────────────────────────────────────────────

def _excel_serial_to_date(serial: float) -> datetime | None:
    try:
        return datetime(1899, 12, 30) + timedelta(days=float(serial))
    except Exception:
        return None


def _parse_date(raw) -> datetime | None:
    """Parse any reasonable date representation into a Python datetime."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw
    if isinstance(raw, pd.Timestamp):
        return raw.to_pydatetime()
    if isinstance(raw, (int, float)):
        return _excel_serial_to_date(raw)

    raw_str = str(raw).strip()
    if not raw_str or raw_str.upper() in ("NONE", "NAT", "NAN", ""):
        return None

    try:
        return _excel_serial_to_date(float(raw_str))
    except ValueError:
        pass

    for fmt in (
        "%d-%b-%y", "%d-%b-%Y",
        "%d/%m/%Y", "%d/%m/%y",
        "%Y-%m-%d", "%m/%d/%Y",
        "%d %b %Y", "%d %B %Y",
        "%B %d, %Y",
    ):
        try:
            return datetime.strptime(raw_str, fmt)
        except ValueError:
            pass

    try:
        return pd.to_datetime(raw_str, dayfirst=True).to_pydatetime()
    except Exception:
        return None


def _format_date_badge(raw) -> tuple[str, str, str]:
    """Return (D-MMM-YY, month_str, year_str) — no zero-padding on day."""
    dt = _parse_date(raw)
    if dt is None:
        return "", "", ""
    return (
        f"{dt.day}-{MONTH_ABBR[dt.month-1]}-{str(dt.year)[-2:]}",
        str(dt.month),
        str(dt.year),
    )


def _format_date_dto(raw) -> str:
    """Return date as  'D/M/YYYY 0:00'  for the DTO template."""
    dt = _parse_date(raw)
    if dt is None:
        return ""
    return f"{dt.day}/{dt.month}/{dt.year} 0:00"


# ── Master file helpers ───────────────────────────────────────────────────────

def _normalise_col(name: str) -> str:
    return re.sub(r"\s+", " ", str(name).strip().lower())


def _map_csv_columns(df: pd.DataFrame) -> dict[str, str]:
    """Map canonical names → actual CSV column names via COLUMN_ALIASES."""
    normalised = {_normalise_col(c): c for c in df.columns}
    mapping: dict[str, str] = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        found = None
        for alias in aliases:
            if alias in normalised:
                found = normalised[alias]
                break
        if found is None:
            for norm_col, real_col in normalised.items():
                if any(alias.split()[0] in norm_col for alias in aliases):
                    found = real_col
                    break
        if found is None:
            raise ValueError(
                f"Cannot find column '{canonical}'. "
                f"Aliases checked: {aliases}. "
                f"Available: {list(df.columns)}"
            )
        mapping[canonical] = found
    return mapping


def _read_master_header(path: str) -> tuple[dict[str, int], dict[str, int]]:
    """
    Read the Masterlist header row.
    Returns (col_map_0based, sheet_col_1based).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[MASTERLIST_SHEET]
    hdr = next(
        ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True),
        None,
    )
    wb.close()
    if not hdr:
        raise ValueError(f"Header row {HEADER_ROW} is empty in {path}")
    col_map   = {str(v).strip(): i for i, v in enumerate(hdr) if v is not None}
    sheet_col = {k: v + 1 for k, v in col_map.items()}
    return col_map, sheet_col


def _read_master_dedup_keys(path: str, col_map: dict[str, int]) -> set[tuple]:
    """Read existing dedup keys from the Masterlist (read-only)."""
    key_positions = [col_map.get(k) for k in DEDUP_KEY]
    existing: set[tuple] = set()
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[MASTERLIST_SHEET]
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        key = tuple(
            str(row[p]).strip().lower()
            if (p is not None and p < len(row) and row[p] is not None) else ""
            for p in key_positions
        )
        existing.add(key)
    wb.close()
    return existing


def _find_last_data_row(ws, anchor_col: int) -> int:
    """Scan from ws.max_row downward; return last row with a value in anchor_col."""
    for r in range(ws.max_row, HEADER_ROW, -1):
        if ws.cell(row=r, column=anchor_col).value is not None:
            return r
    return HEADER_ROW


def _append_rows_to_master(new_rows: pd.DataFrame, sheet_col: dict[str, int]) -> None:
    """Append new_rows to the Masterlist sheet without touching anything else."""
    centre_cols = {sheet_col[f] for f in ("Date of Award", "Month", "Year") if f in sheet_col}
    anchor_col  = sheet_col.get("Name") or sheet_col.get("Email") or min(sheet_col.values())

    wb = load_workbook(MASTER_FILE)
    ws = wb[MASTERLIST_SHEET]
    next_row = _find_last_data_row(ws, anchor_col) + 1

    logger.debug("  Master append: anchor_col=%d  next_row=%d", anchor_col, next_row)

    for _, data_row in new_rows.iterrows():
        for field, value in data_row.items():
            col_num = sheet_col.get(str(field).strip())
            if col_num is None:
                continue
            cell = ws.cell(row=next_row, column=col_num,
                           value=value if value != "" else None)
            if col_num in centre_cols:
                cell.alignment = Alignment(horizontal="center")
        next_row += 1

    wb.save(MASTER_FILE)


def _move_to_processed(src: str) -> None:
    filename = os.path.basename(src)
    dest = os.path.join(PROCESSED_FOLDER, filename)
    if os.path.exists(dest):
        stem, ext = os.path.splitext(filename)
        dest = os.path.join(PROCESSED_FOLDER,
                            f"{stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}")
    shutil.move(src, dest)
    logger.info("  Moved → processed: %s", os.path.basename(dest))


# =============================================================================
# PART 1 — TRANSFORM CSV → MASTERLIST
# =============================================================================

def _transform_csv(filepath: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Read one CSV, transform it, and return:
      (master_rows_df, dto_rows_df)

    master_rows_df  — WRITE_COLUMNS format ready for the Masterlist
    dto_rows_df     — original CSV column format for the combined DTO CSV export
    """
    try:
        df = pd.read_csv(filepath, encoding="utf-8-sig", dtype=str)
    except UnicodeDecodeError:
        df = pd.read_csv(filepath, encoding="latin-1", dtype=str)
    df.columns = [c.strip() for c in df.columns]

    col = _map_csv_columns(df)

    master_records = []
    dto_records    = []

    for _, row in df.iterrows():
        email     = str(row[col["Email"]]).strip().upper()
        name      = str(row[col["Preferred Name"]]).strip()
        badge_raw = str(row[col["Skills Badge"]]).strip()
        programme = str(row[col["Programme"]]).strip()
        date_raw  = row[col["Date"]]
        provider  = str(row[col["Training Provider"]]).strip()

        if not email or not name or not badge_raw or email == "NAN":
            continue

        # Parse badge name into parts
        match = re.match(r"^(.*?)\s*\(([^)]+)\)\s*$", badge_raw)
        if match:
            skills_area       = match.group(1).strip()
            skills_area_level = badge_raw
            badge_level       = match.group(2).strip()
        else:
            skills_area       = badge_raw
            skills_area_level = badge_raw
            badge_level       = ""

        # Fallback: look for a "badge level" column in the CSV
        if not badge_level:
            for col_name in df.columns:
                if "level" in col_name.lower() and "badge" in col_name.lower():
                    fallback = str(row[col_name]).strip()
                    if fallback and fallback.upper() != "NAN":
                        badge_level       = fallback
                        skills_area_level = f"{skills_area} ({badge_level})"
                    break

        date_display, month, year = _format_date_badge(date_raw)

        master_records.append({
            "Name":                  name,
            "Email":                 email,
            "Training Provider":     provider,
            "Skills Area":           skills_area,
            "Skills Area and Level": skills_area_level,
            "Badge Level":           badge_level,
            "Date of Award":         date_display,
            "Month":                 month,
            "Year":                  year,
            "Programme":             programme,
        })

        # Keep a copy in the original CSV column names for the DTO CSV export
        dto_records.append({
            "Identifier (Email_Address)":          email,
            "Preferred Name(Name to appear on badge)": name,
            "Name of Skills Badge":                skills_area_level,
            "Course / Programme Title":            programme,
            "Date of Course Completion":           date_display,
            "Training Provider":                   provider,
        })

    master_df = pd.DataFrame(master_records, columns=WRITE_COLUMNS)
    dto_df    = pd.DataFrame(dto_records)
    return master_df, dto_df


def run_part1() -> tuple[list[pd.DataFrame], list[pd.DataFrame]]:
    """
    PART 1: process all incoming CSVs.
    Returns (all_master_new_rows, all_dto_rows) — one DataFrame per file.
    """
    logger.info("=" * 60)
    logger.info("PART 1 — Transform CSVs → Masterlist")
    logger.info("=" * 60)

    csv_files = sorted(glob.glob(os.path.join(INPUT_FOLDER, "*.csv")))
    if not csv_files:
        logger.info("No CSV files found in %s", INPUT_FOLDER)
        return [], []

    os.makedirs(PROCESSED_FOLDER, exist_ok=True)

    col_map, sheet_col = _read_master_header(MASTER_FILE)
    existing_keys      = _read_master_dedup_keys(MASTER_FILE, col_map)

    def _key(row) -> tuple:
        def _v(k):
            try:
                v = row[k]
                return str(v).strip().lower() if v is not None and str(v).upper() != "NAN" else ""
            except KeyError:
                return ""
        return tuple(_v(k) for k in DEDUP_KEY)

    all_master_rows: list[pd.DataFrame] = []
    all_dto_rows:    list[pd.DataFrame] = []

    for filepath in csv_files:
        filename  = os.path.basename(filepath)
        timestamp = _ts()
        logger.info("Processing CSV: %s", filename)

        try:
            master_df, dto_df = _transform_csv(filepath)

            if master_df.empty:
                logger.warning("  No valid rows in %s", filename)
                _append_log(timestamp, "Part1", filename, 0, "SKIPPED", "No valid rows")
                _move_to_processed(filepath)
                continue

            # Deduplicate against master
            new_rows = master_df[
                master_df.apply(_key, axis=1).apply(lambda k: k not in existing_keys)
            ].copy()
            dup_count = len(master_df) - len(new_rows)
            if dup_count:
                logger.info("  Skipped %d duplicate(s).", dup_count)

            if new_rows.empty:
                logger.info("  All rows already in master.")
                _append_log(timestamp, "Part1", filename, 0, "SKIPPED",
                            "All rows already in master")
                _move_to_processed(filepath)
                continue

            _append_rows_to_master(new_rows[WRITE_COLUMNS], sheet_col)

            for _, r in new_rows.iterrows():
                existing_keys.add(_key(r))

            # Keep only the DTO rows that correspond to new (non-duplicate) master rows
            # Match on email + badge name since dto_df has no dedup key
            new_emails_badges = set(
                zip(new_rows["Email"].str.lower(), new_rows["Skills Area and Level"].str.lower())
            )
            new_dto = dto_df[
                dto_df.apply(
                    lambda r: (
                        str(r.get("Identifier (Email_Address)", "")).lower(),
                        str(r.get("Name of Skills Badge", "")).lower(),
                    ) in new_emails_badges,
                    axis=1,
                )
            ].copy()

            all_master_rows.append(new_rows)
            all_dto_rows.append(new_dto)

            rows_added = len(new_rows)
            logger.info("  Appended %d row(s) to master.", rows_added)
            _append_log(timestamp, "Part1", filename, rows_added, "SUCCESS",
                        f"{dup_count} duplicate(s) skipped")
            _move_to_processed(filepath)

        except Exception as exc:
            logger.error("  FAILED: %s", exc)
            logger.debug(traceback.format_exc())
            _append_log(_ts(), "Part1", filename, 0, "ERROR",
                        f"{type(exc).__name__}: {exc}")

    return all_master_rows, all_dto_rows


# =============================================================================
# PART 2 — AWARD ROLE BADGES
# =============================================================================

def _build_user_data_for_roles(col_map: dict[str, int]) -> dict:
    """Read the current Masterlist (including Part 1 additions) into a user dict."""
    c_email = col_map["Email"]
    c_name  = col_map["Name"]
    c_area  = col_map["Skills Area"]
    c_level = col_map["Skills Area and Level"]
    c_badge = col_map["Badge Level"]
    c_date  = col_map["Date of Award"]

    wb = load_workbook(MASTER_FILE, read_only=True, data_only=True)
    ws = wb[MASTERLIST_SHEET]
    rows = list(ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True))
    wb.close()

    users: dict = {}
    for row in rows:
        if all(v is None for v in row):
            continue

        def _g(c):
            v = row[c] if c < len(row) else None
            return str(v).strip() if v is not None else ""

        email_key = _g(c_email).lower()
        if not email_key:
            continue

        name           = _g(c_name)
        skills_area    = _g(c_area)
        area_and_level = _g(c_level)
        badge_level    = _g(c_badge)
        date_raw       = row[c_date] if c_date < len(row) else None
        date_obj       = _parse_date(date_raw)

        if email_key not in users:
            users[email_key] = {
                "name":                 name,
                "email_raw":            _g(c_email),
                "skills":               set(),
                "badge_rows":           [],
                "existing_role_badges": set(),
            }
        elif not users[email_key]["name"] and name:
            users[email_key]["name"] = name

        if skills_area:
            users[email_key]["skills"].add(skills_area.lower())

        if badge_level.lower() != "role badge" and skills_area:
            users[email_key]["badge_rows"].append({
                "skills_area":       skills_area,
                "skills_area_level": area_and_level,
                "date_obj":          date_obj,
            })

        if badge_level.lower() == "role badge" and area_and_level:
            users[email_key]["existing_role_badges"].add(area_and_level.lower())

    return users


def _natural_list(items: list[str]) -> str:
    if len(items) == 1:
        return items[0]
    return ", ".join(items[:-1]) + " and " + items[-1]


def _derive_role_badge_display(role_badge_name: str) -> tuple[str, str, str]:
    """
    Given a dict-key role badge name, return:
      (skills_area_clean, skills_area_and_level, level_tag)
    e.g. "Workplace Learning Champion Role Badge"
      → ("Workplace Learning Champion",
         "Workplace Learning Champion (Role Badge)",
         "Role Badge")
    """
    clean = role_badge_name
    for sfx in (" Role Badge (L1)", " Role Badge (L2)", " Role Badge"):
        if clean.lower().endswith(sfx.lower()):
            clean = clean[: -len(sfx)].strip()
            break
    if "(L1)" in role_badge_name:
        tag = "Role Badge (L1)"
    elif "(L2)" in role_badge_name:
        tag = "Role Badge (L2)"
    else:
        tag = "Role Badge"
    return clean, f"{clean} ({tag})", tag


def run_part2(sheet_col: dict[str, int]) -> list[pd.DataFrame]:
    """
    PART 2: check for newly eligible role badges and append them.
    Returns list of DataFrames (one per role badge) with new rows in WRITE_COLUMNS format.
    """
    logger.info("=" * 60)
    logger.info("PART 2 — Award Role Badges")
    logger.info("=" * 60)

    col_map, sheet_col = _read_master_header(MASTER_FILE)
    users = _build_user_data_for_roles(col_map)
    logger.info("Loaded %d unique users.", len(users))

    centre_cols = {sheet_col[f] for f in ("Date of Award", "Month", "Year") if f in sheet_col}
    anchor_col  = sheet_col.get("Name") or sheet_col.get("Email") or min(sheet_col.values())

    all_role_rows: list[pd.DataFrame] = []

    for role_badge_name, required_skills in ROLE_BADGE_REQUIREMENTS.items():
        required_lower      = {s.lower() for s in required_skills}
        role_badge_lower    = role_badge_name.lower()
        skills_area_clean, skills_area_and_level, level_tag = \
            _derive_role_badge_display(role_badge_name)
        bracket_lower = skills_area_and_level.lower()

        logger.info("Checking: %s", role_badge_name)

        eligible = [
            u for u in users.values()
            if required_lower.issubset(u["skills"])
            and role_badge_lower not in u["existing_role_badges"]
            and bracket_lower   not in u["existing_role_badges"]
        ]

        if not eligible:
            logger.info("  No new recipients.")
            continue

        logger.info("  %d new recipient(s).", len(eligible))

        wb       = load_workbook(MASTER_FILE)
        ws       = wb[MASTERLIST_SHEET]
        next_row = _find_last_data_row(ws, anchor_col) + 1
        records  = []

        for user in eligible:
            email     = user["email_raw"]
            name      = user["name"]
            timestamp = _ts()

            qualifying = [
                br for br in user["badge_rows"]
                if br["skills_area"].lower() in required_lower
            ]
            date_objs = [br["date_obj"] for br in qualifying if br["date_obj"]]
            latest_dt = max(date_objs) if date_objs else None
            if latest_dt is None:
                logger.warning("    No date found for %s — qualifying: %s",
                               email,
                               [(b["skills_area"], b.get("date_obj")) for b in qualifying])
            date_display, month_str, year_str = (
                _format_date_badge(latest_dt) if latest_dt else ("", "", "")
            )

            # Build programme label
            badge_labels, seen = [], set()
            for req in required_skills:
                matches = sorted(
                    [b for b in qualifying if b["skills_area"].lower() == req.lower()],
                    key=lambda x: x["date_obj"] or datetime.min,
                    reverse=True,
                )
                if matches:
                    lbl = matches[0]["skills_area_level"] or matches[0]["skills_area"]
                    if lbl not in seen:
                        badge_labels.append(lbl)
                        seen.add(lbl)
            programme = f"Attainment of Skills Badges: {_natural_list(badge_labels)}"

            new_data = {
                "Name":                  name,
                "Email":                 email,
                "Training Provider":     ROLE_BADGE_PROVIDER,
                "Skills Area":           skills_area_clean,
                "Skills Area and Level": skills_area_and_level,
                "Badge Level":           "Role Badge",
                "Date of Award":         date_display,
                "Month":                 month_str,
                "Year":                  year_str,
                "Programme":             programme,
            }

            # Write to sheet
            for field, value in new_data.items():
                col_num = sheet_col.get(field)
                if col_num is None:
                    continue
                cell = ws.cell(row=next_row, column=col_num,
                               value=value if value != "" else None)
                if col_num in centre_cols:
                    cell.alignment = Alignment(horizontal="center")

            records.append(new_data)
            logger.info("    + %s → %s (%s)", email, skills_area_and_level, date_display)
            _append_log(timestamp, "Part2", email, 1, "AWARDED",
                        f"{skills_area_and_level} | {date_display} | {programme}")

            users[email.lower()]["existing_role_badges"].add(role_badge_lower)
            users[email.lower()]["existing_role_badges"].add(bracket_lower)
            next_row += 1

        wb.save(MASTER_FILE)

        if records:
            all_role_rows.append(pd.DataFrame(records, columns=WRITE_COLUMNS))

    return all_role_rows


# =============================================================================
# PART 3 — EXPORT TO DTO FOLDER
# =============================================================================

DTO_TEMPLATE_COLUMNS = [
    "ID", "Document Name", "Issued On", "Expires On",
    "Admission Date", "Graduation Date", "Attainment Date",
    "Recipient Name", "Recipient Email", "Recipient NRIC",
    "Recipient Student ID", "Training Provider", "Programme",
]


def _master_row_to_dto(row: dict) -> dict:
    """Convert one WRITE_COLUMNS row dict to a DTO template row dict."""
    return {
        "ID":                   "",
        "Document Name":        row.get("Skills Area and Level", ""),
        "Issued On":            _format_date_dto(row.get("Date of Award", "")),
        "Expires On":           "",
        "Admission Date":       "",
        "Graduation Date":      "",
        "Attainment Date":      "",
        "Recipient Name":       row.get("Name", ""),
        "Recipient Email":      row.get("Email", ""),
        "Recipient NRIC":       "",
        "Recipient Student ID": "",
        "Training Provider":    row.get("Training Provider", ""),
        "Programme":            row.get("Programme", ""),
    }


def _safe_filename(name: str) -> str:
    """Strip characters that are illegal in Windows filenames."""
    return re.sub(r'[\/:*?"<>|]', "_", name).strip()


def _write_dto_excel(rows: list[dict], out_folder: str,
                     stamp: str, label: str) -> str:
    """
    Copy the template, fill it with *rows* (list of WRITE_COLUMNS dicts),
    save to out_folder, and return the saved path.
    *label* is used in the filename (e.g. "All" or a badge name).
    """
    safe_label = _safe_filename(label)
    dest = os.path.join(out_folder, f"Badges_DTO_{safe_label}_{stamp}.xlsx")
    shutil.copy2(BADGE_TEMPLATE_FILE, dest)

    wb  = load_workbook(dest)
    ws  = wb.active

    # Build header map from template row 3
    hdr = {
        str(cell.value).strip(): cell.column
        for cell in ws[3]
        if cell.value is not None
    }

    # Find first empty data row (anchor on "Document Name" col)
    doc_col  = hdr.get("Document Name", 2)
    next_row = 4  # default: row 4 is first data row
    for r in range(ws.max_row, 3, -1):
        if ws.cell(row=r, column=doc_col).value is not None:
            next_row = r + 1
            break

    for row_dict in rows:
        dto = _master_row_to_dto(row_dict)
        for col_name, col_num in hdr.items():
            val = dto.get(col_name, "")
            ws.cell(row=next_row, column=col_num,
                    value=val if val != "" else None)
        next_row += 1

    wb.save(dest)
    return dest


def _write_combined_csv(rows: list[dict], out_folder: str,
                        stamp: str, label: str) -> str:
    """Write rows as a CSV in the original incoming-CSV column format."""
    safe_label = _safe_filename(label)
    dest = os.path.join(out_folder, f"Combined_{safe_label}_{stamp}.csv")

    CSV_COLUMNS = [
        "Identifier (Email_Address)",
        "Preferred Name(Name to appear on badge)",
        "Name of Skills Badge",
        "Badge Level",
        "Course / Programme Title",
        "Date of Course Completion",
        "Training Provider",
    ]
    records = []
    for r in rows:
        records.append({
            "Identifier (Email_Address)":              r.get("Email", ""),
            "Preferred Name(Name to appear on badge)": r.get("Name", ""),
            "Name of Skills Badge":                    r.get("Skills Area and Level", ""),
            "Badge Level":                             r.get("Badge Level", ""),
            "Course / Programme Title":                r.get("Programme", ""),
            "Date of Course Completion":               r.get("Date of Award", ""),
            "Training Provider":                       r.get("Training Provider", ""),
        })
    pd.DataFrame(records, columns=CSV_COLUMNS).to_csv(
        dest, index=False, encoding="utf-8-sig"
    )
    return dest


def run_part3(
    part1_master_rows: list[pd.DataFrame],
    part2_role_rows:   list[pd.DataFrame],
) -> None:
    """
    PART 3 — Export to DTO folder.

    Respects PROCESS_MODE (which rows to include) and EXPORT_MODE (how to split).

    PROCESS_MODE:
      "skills" → export only Part 1 (skill badge) rows
      "roles"  → export only Part 2 (role badge) rows
      "both"   → export all rows

    EXPORT_MODE:
      "combined" → one Excel + one CSV with all selected rows
      "split"    → one Excel + one CSV per unique "Skills Area and Level" value
    """
    logger.info("=" * 60)
    logger.info("PART 3 — Export to DTO folder  (process=%s  export=%s)",
                PROCESS_MODE, EXPORT_MODE)
    logger.info("=" * 60)

    # Select which rows to export based on PROCESS_MODE
    if PROCESS_MODE == "skills":
        export_dfs = part1_master_rows
    elif PROCESS_MODE == "roles":
        export_dfs = part2_role_rows
    else:  # "both"
        export_dfs = part1_master_rows + part2_role_rows

    # Flatten to a list of plain dicts
    all_rows: list[dict] = []
    for df in export_dfs:
        for _, r in df.iterrows():
            all_rows.append(dict(r))

    if not all_rows:
        logger.info("No new rows to export — skipping Part 3.")
        _append_log(_ts(), "Part3", "export", 0, "SKIPPED", "No new rows")
        return

    # Create timestamped output folder
    stamp      = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_folder = os.path.join(TO_DTO_BASE_FOLDER, stamp)
    os.makedirs(out_folder, exist_ok=True)
    logger.info("Output folder: %s", out_folder)

    if EXPORT_MODE == "combined":
        # ── One Excel + one CSV with everything ───────────────────────────
        xlsx_path = _write_dto_excel(all_rows, out_folder, stamp, "All")
        csv_path  = _write_combined_csv(all_rows, out_folder, stamp, "All")

        logger.info("Excel written: %s (%d rows)", os.path.basename(xlsx_path), len(all_rows))
        logger.info("CSV   written: %s (%d rows)", os.path.basename(csv_path),  len(all_rows))
        _append_log(_ts(), "Part3", os.path.basename(xlsx_path),
                    len(all_rows), "SUCCESS",
                    f"mode=combined | {out_folder}")
        _append_log(_ts(), "Part3", os.path.basename(csv_path),
                    len(all_rows), "SUCCESS",
                    f"mode=combined | {out_folder}")

    else:  # "split"
        # ── One Excel + one CSV per unique badge name ──────────────────────
        # Group rows by "Skills Area and Level"
        from collections import defaultdict
        groups: dict[str, list[dict]] = defaultdict(list)
        for row in all_rows:
            badge_name = row.get("Skills Area and Level", "Unknown Badge")
            groups[badge_name].append(row)

        logger.info("Split mode: %d unique badge(s) → %d file pair(s)",
                    len(groups), len(groups))

        for badge_name, badge_rows in sorted(groups.items()):
            xlsx_path = _write_dto_excel(badge_rows, out_folder, stamp, badge_name)
            csv_path  = _write_combined_csv(badge_rows, out_folder, stamp, badge_name)

            logger.info("  [%s] Excel: %s (%d rows)",
                        badge_name, os.path.basename(xlsx_path), len(badge_rows))
            logger.info("  [%s] CSV:   %s (%d rows)",
                        badge_name, os.path.basename(csv_path),  len(badge_rows))
            _append_log(_ts(), "Part3", os.path.basename(xlsx_path),
                        len(badge_rows), "SUCCESS",
                        f"mode=split | badge={badge_name} | {out_folder}")
            _append_log(_ts(), "Part3", os.path.basename(csv_path),
                        len(badge_rows), "SUCCESS",
                        f"mode=split | badge={badge_name} | {out_folder}")
# =============================================================================
# MAIN PIPELINE
# =============================================================================

def run_pipeline() -> None:
    valid_process = ("skills", "roles", "both")
    valid_export  = ("combined", "split")
    if PROCESS_MODE not in valid_process:
        raise ValueError(
            f"PROCESS_MODE must be one of {valid_process}, got {PROCESS_MODE!r}"
        )
    if EXPORT_MODE not in valid_export:
        raise ValueError(
            f"EXPORT_MODE must be one of {valid_export}, got {EXPORT_MODE!r}"
        )

    logger.info("╔══════════════════════════════════════════════════════════╗")
    logger.info("║  BADGE PIPELINE — START  (process=%s  export=%s)",
                PROCESS_MODE, EXPORT_MODE)
    logger.info("╚══════════════════════════════════════════════════════════╝")

    _, sheet_col = _read_master_header(MASTER_FILE)

    # ── Part 1: import CSVs (only when mode requires it) ──────────────────────
    part1_master_rows: list[pd.DataFrame] = []
    if PROCESS_MODE in ("skills", "both"):
        part1_master_rows, _ = run_part1()
    else:
        logger.info("PROCESS_MODE=%r — skipping Part 1 (CSV import).", PROCESS_MODE)

    # ── Part 2: award role badges (only when mode requires it) ────────────────
    part2_role_rows: list[pd.DataFrame] = []
    if PROCESS_MODE in ("roles", "both"):
        part2_role_rows = run_part2(sheet_col)
    else:
        logger.info("PROCESS_MODE=%r — skipping Part 2 (role badge check).", PROCESS_MODE)

    # ── Part 3: export to To_DTO folder ───────────────────────────────────────
    run_part3(part1_master_rows, part2_role_rows)

    logger.info("╔══════════════════════════════════════════════════════════╗")
    logger.info("║              BADGE PIPELINE — DONE                      ║")
    logger.info("╚══════════════════════════════════════════════════════════╝")


if __name__ == "__main__":
    run_pipeline()
