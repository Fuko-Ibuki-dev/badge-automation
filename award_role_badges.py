"""
award_role_badges.py
=====================
Checks the Masterlist sheet for users who have earned all required skill
badges for a role badge but have not yet been awarded that role badge.
For each qualifying user, one new role-badge row is appended to the
Masterlist and the action is recorded in an Excel log file.
 
USAGE
-----
1. Edit the CONFIG section below (same paths as transform_badges.py).
2. Run:  python award_role_badges.py
 
HOW IT WORKS
------------
For each role badge:
  1. Find all users whose "Skills Area" column contains EVERY required skill.
  2. Check whether that user already has a row whose "Skills Area and Level"
     matches the role badge name (case-insensitive).
  3. If not, append a new row:
       • Name / Email         — copied from the user's existing rows
       • Training Provider    — "Company"
       • Skills Area          — role badge name (without " Role Badge" suffix)
       • Skills Area and Level— full role badge name
       • Badge Level          — "Role Badge"
       • Date of Award        — latest "Date of Award" among the qualifying rows
       • Month / Year         — derived from Date of Award
       • Programme            — "Attainment of Skills Badges: X, Y and Z"
 
ASSUMPTIONS
-----------
- The master file and log file paths are shared with transform_badges.py.
- Header is on HEADER_ROW (1-based) of the MASTERLIST_SHEET sheet.
- A user is identified by their (lowercased, stripped) Email address.
- "Date of Award" values are in "DD-MMM-YY" format (as written by
  transform_badges.py).  Other parseable formats also work.
"""

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG  ← must match the paths used in transform_badges.py
# ─────────────────────────────────────────────────────────────────────────────
 
MASTER_FILE      = r"C:\Users\GaoMing\Desktop\Python\badge_automation\mock_master_list.xlsx"
LOG_FILE         = r"C:\Users\GaoMing\Desktop\Python\badge_automation\role_badge_log.xlsx"
MASTERLIST_SHEET = "Masterlist"
HEADER_ROW       = 3          # 1-based row number of the header in Masterlist
ROLE_BADGE_PROVIDER = "Company"  # Training Provider written on every role badge row
 
# ─────────────────────────────────────────────────────────────────────────────
# ROLE BADGE → REQUIRED SKILLS MAPPING
# Keys   : exact role badge name (used in "Skills Area and Level" column)
# Values : list of required "Skills Area" values (all must be present)
# ─────────────────────────────────────────────────────────────────────────────
 
ROLE_BADGE_REQUIREMENTS: dict[str, list[str]] = {
    "Example Role Badge": [
        "Skills Badge A",
        "Skills Badge B",
        "Skills Badge C",
    ],
}
 
# ─────────────────────────────────────────────────────────────────────────────
# IMPORTS
# ─────────────────────────────────────────────────────────────────────────────
 
import os
import logging
from datetime import datetime, timedelta
from collections import defaultdict
 
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
 
# ─────────────────────────────────────────────────────────────────────────────
# LOGGING SETUP
# ─────────────────────────────────────────────────────────────────────────────
 
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)
 
# ─────────────────────────────────────────────────────────────────────────────
# DATE HELPERS  (mirror the logic in transform_badges.py)
# ─────────────────────────────────────────────────────────────────────────────
 
MONTH_ABBR = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]
 
 
def _parse_date(raw) -> datetime | None:
    """Parse a date value (string, datetime, or Excel serial) into a Python datetime."""
    if raw is None:
        return None
 
    # openpyxl with data_only=True returns date cells as Python datetime objects
    if isinstance(raw, datetime):
        return raw
 
    raw_str = str(raw).strip()
    if not raw_str or raw_str.upper() in ("NONE", "NAT", "NAT", ""):
        return None
 
    # Excel serial number stored as string
    try:
        serial = float(raw_str)
        if serial > 0:
            return datetime(1899, 12, 30) + timedelta(days=serial)
    except ValueError:
        pass
 
    for fmt in (
        "%d-%b-%y",   # 22-Aug-25
        "%d-%b-%Y",   # 22-Aug-2025
        "%d/%m/%Y",
        "%d/%m/%y",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%d %b %Y",
        "%d %B %Y",
        "%B %d, %Y",
    ):
        try:
            return datetime.strptime(raw_str, fmt)
        except ValueError:
            pass
    return None
 
 
def _format_date(dt: datetime) -> tuple[str, str, str]:
    """Return (DD-MMM-YY, month_int_str, year_str) for a datetime."""
    dd  = str(dt.day).zfill(2)
    mmm = MONTH_ABBR[dt.month - 1]
    yy  = str(dt.year)[-2:]
    return f"{dd}-{mmm}-{yy}", str(dt.month), str(dt.year)
 
 
# ─────────────────────────────────────────────────────────────────────────────
# EXCEL LOG HELPER
# ─────────────────────────────────────────────────────────────────────────────
 
def _append_log(file: str, timestamp: str, email: str, role_badge: str,
                status: str, detail: str = "") -> None:
    """Append one row to the Excel log file, creating it if needed."""
    LOG_HEADERS = ["Timestamp", "Email", "Role Badge", "Status", "Detail"]
 
    if os.path.isfile(file):
        wb = load_workbook(file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "RoleBadgeLog"
        ws.append(LOG_HEADERS)
 
        fill = PatternFill("solid", fgColor="BDD7EE")
        for cell in ws[1]:
            cell.font      = Font(bold=True)
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center")
 
        ws.freeze_panes = "A2"
        for col_letter, width in zip("ABCDE", [22, 40, 52, 12, 60]):
            ws.column_dimensions[col_letter].width = width
 
    ws.append([timestamp, email, role_badge, status, detail])
    wb.save(file)
 
 
# ─────────────────────────────────────────────────────────────────────────────
# READ MASTERLIST
# ─────────────────────────────────────────────────────────────────────────────
 
def _read_masterlist(path: str) -> tuple[dict, list, dict]:
    """
    Load the Masterlist sheet into memory.
 
    Returns
    -------
    col_map : dict {header_name: 0-based index}
    rows    : list of tuples, one per data row (values only, no header)
    sheet_col : dict {header_name: 1-based Excel column number}
        Needed later when we write new rows back via openpyxl.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
 
    if MASTERLIST_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f'Sheet "{MASTERLIST_SHEET}" not found. '
            f"Available: {wb.sheetnames}"
        )
 
    ws = wb[MASTERLIST_SHEET]
 
    # Read header row
    header_tuple = next(
        ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True),
        None,
    )
    if not header_tuple:
        wb.close()
        raise ValueError(f"Header row {HEADER_ROW} is empty in {path}")
 
    # 0-based index map (for reading rows as tuples)
    col_map = {
        str(v).strip(): i
        for i, v in enumerate(header_tuple)
        if v is not None
    }
 
    # 1-based column map (for writing cells back with openpyxl)
    sheet_col = {k: v + 1 for k, v in col_map.items()}
 
    # Read all data rows
    rows = list(
        ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True)
    )
 
    wb.close()
    return col_map, rows, sheet_col
 
 
# ─────────────────────────────────────────────────────────────────────────────
# CORE LOGIC
# ─────────────────────────────────────────────────────────────────────────────
 
def _build_user_data(col_map: dict, rows: list) -> dict:
    """
    Build a per-user data structure from the raw masterlist rows.
 
    Returns
    -------
    users : dict  {email_lower: {
                    "name"         : str,
                    "email_raw"    : str,   # original casing
                    "skills"       : set of lowercased Skills Area strings,
                    "badge_rows"   : list of {"skills_area": str,
                                              "skills_area_level": str,
                                              "date_obj": datetime|None,
                                              "date_str": str},
                    "existing_role_badges": set of lowercased role badge names,
                  }}
    """
    # Required column positions
    c_email   = col_map.get("Email")
    c_name    = col_map.get("Name")
    c_area    = col_map.get("Skills Area")
    c_level   = col_map.get("Skills Area and Level")
    c_badge   = col_map.get("Badge Level")
    c_date    = col_map.get("Date of Award")
 
    missing = [
        name for name, pos in [
            ("Email", c_email), ("Name", c_name),
            ("Skills Area", c_area), ("Skills Area and Level", c_level),
            ("Badge Level", c_badge), ("Date of Award", c_date),
        ] if pos is None
    ]
    if missing:
        raise ValueError(f"Required columns not found in master header: {missing}")
 
    users: dict = {}
 
    for row in rows:
        # Skip fully empty rows
        if all(v is None for v in row):
            continue
 
        def _get(col):
            v = row[col] if col < len(row) else None
            return str(v).strip() if v is not None else ""
 
        email_raw  = _get(c_email)
        email_key  = email_raw.lower()
        if not email_key:
            continue
 
        name           = _get(c_name)
        skills_area    = _get(c_area)
        area_and_level = _get(c_level)
        badge_level    = _get(c_badge)
        date_raw       = row[c_date] if c_date < len(row) else None
        date_obj       = _parse_date(date_raw)
        date_str       = _get(c_date)
 
        if email_key not in users:
            users[email_key] = {
                "name":                 name,
                "email_raw":            email_raw,
                "skills":               set(),
                "badge_rows":           [],
                "existing_role_badges": set(),
            }
        elif not users[email_key]["name"] and name:
            users[email_key]["name"] = name
 
        if skills_area:
            users[email_key]["skills"].add(skills_area.lower())
 
        # Track non-role-badge rows (for date lookup)
        if badge_level.lower() != "role badge" and skills_area:
            users[email_key]["badge_rows"].append({
                "skills_area":       skills_area,
                "skills_area_level": area_and_level,
                "date_obj":          date_obj,
                "date_str":          date_str,
            })
 
        # Track already-awarded role badges
        if badge_level.lower() == "role badge" and area_and_level:
            users[email_key]["existing_role_badges"].add(area_and_level.lower())
 
    return users
 
 
def _natural_list(items: list[str]) -> str:
    """
    Join a list into natural English:  "A, B and C"  or  "A and B"  or  "A".
    """
    if len(items) == 1:
        return items[0]
    return ", ".join(items[:-1]) + " and " + items[-1]
 
 
def _check_and_award(path: str) -> None:
    """
    Main logic: read the masterlist, find eligible users, append role badge rows.
    """
    logger.info("Reading masterlist: %s", path)
    col_map, rows, sheet_col = _read_masterlist(path)
    users = _build_user_data(col_map, rows)
 
    logger.info("Loaded %d unique users.", len(users))
 
    # Columns we need for writing; centre-aligned ones
    centre_fields = {"Date of Award", "Month", "Year"}
    centre_cols   = {sheet_col[f] for f in centre_fields if f in sheet_col}
 
    # Anchor column for finding the last data row (same logic as transform_badges.py)
    anchor_col = (
        sheet_col.get("Name") or
        sheet_col.get("Email") or
        min(sheet_col.values())
    )
 
    new_rows_added = 0
 
    for role_badge_name, required_skills in ROLE_BADGE_REQUIREMENTS.items():
        required_lower = {s.lower() for s in required_skills}
        role_badge_lower = role_badge_name.lower()
 
        logger.info("Checking: %s", role_badge_name)
 
        # Derive the bracket-style name for this role badge (used in the sheet)
        _skills_area_tmp = role_badge_name
        for _sfx in (" Role Badge (L1)", " Role Badge (L2)", " Role Badge"):
            if _skills_area_tmp.lower().endswith(_sfx.lower()):
                _skills_area_tmp = _skills_area_tmp[:-len(_sfx)].strip()
                break
        if "(L1)" in role_badge_name:
            _lvl_tag = "Role Badge (L1)"
        elif "(L2)" in role_badge_name:
            _lvl_tag = "Role Badge (L2)"
        else:
            _lvl_tag = "Role Badge"
        role_badge_bracket_lower = f"{_skills_area_tmp} ({_lvl_tag})".lower()
 
        eligible_users = [
            u for u in users.values()
            if required_lower.issubset(u["skills"])
            and role_badge_lower not in u["existing_role_badges"]
            and role_badge_bracket_lower not in u["existing_role_badges"]
        ]
 
        if not eligible_users:
            logger.info("  No new recipients.")
            continue
 
        logger.info("  %d new recipient(s) found.", len(eligible_users))
 
        # Open workbook once per role badge (we need write mode)
        wb = load_workbook(path)
        ws = wb[MASTERLIST_SHEET]
 
        # Find the last occupied row using the anchor column
        last_data_row = HEADER_ROW
        for r in range(ws.max_row, HEADER_ROW, -1):
            if ws.cell(row=r, column=anchor_col).value is not None:
                last_data_row = r
                break
        next_row = last_data_row + 1
 
        for user in eligible_users:
            email     = user["email_raw"]
            name      = user["name"]
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
 
            # ── Find the qualifying skill badge rows for this role badge ──────
            # We only look at rows whose Skills Area matches one of the required skills
            qualifying = [
                br for br in user["badge_rows"]
                if br["skills_area"].lower() in required_lower
            ]
 
            # Latest date among qualifying rows
            date_objs = [br["date_obj"] for br in qualifying if br["date_obj"]]
            latest_dt = max(date_objs) if date_objs else None
            if latest_dt is None:
                logger.warning(
                    "    Could not determine Date of Award for %s — "
                    "qualifying rows: %s",
                    email,
                    [(br["skills_area"], repr(br.get("date_str"))) for br in qualifying],
                )
            date_display, month_str, year_str = (
                _format_date(latest_dt) if latest_dt else ("", "", "")
            )
 
            # Build "Skills Area and Level" labels for the Programme field
            # Use the full "Skills Area and Level" value from each qualifying row,
            # falling back to just the Skills Area if that column is blank.
            badge_labels = []
            seen_areas   = set()
            for req_skill in required_skills:
                # Find the most recent qualifying row for this skill
                matches = sorted(
                    [br for br in qualifying if br["skills_area"].lower() == req_skill.lower()],
                    key=lambda x: x["date_obj"] or datetime.min,
                    reverse=True,
                )
                if matches:
                    label = matches[0]["skills_area_level"] or matches[0]["skills_area"]
                    if label not in seen_areas:
                        badge_labels.append(label)
                        seen_areas.add(label)
 
            programme = f"Attainment of Skills Badges: {_natural_list(badge_labels)}"
 
            # Skills Area = role badge name with " Role Badge" suffix removed.
            # Skills Area and Level = same but with "(Role Badge)" appended in
            # bracket style to match the convention used for skill badges,
            # e.g. "Workplace Learning Champion (Role Badge)".
            skills_area_clean = role_badge_name
            for suffix in (" Role Badge (L1)", " Role Badge (L2)", " Role Badge"):
                if skills_area_clean.lower().endswith(suffix.lower()):
                    skills_area_clean = skills_area_clean[:-len(suffix)].strip()
                    break
 
            # Derive the level tag from the original name for bracket notation
            if "(L1)" in role_badge_name:
                level_tag = "Role Badge (L1)"
            elif "(L2)" in role_badge_name:
                level_tag = "Role Badge (L2)"
            else:
                level_tag = "Role Badge"
 
            skills_area_and_level = f"{skills_area_clean} ({level_tag})"
 
            # ── Write the new row to the sheet ────────────────────────────────
            new_data = {
                "Name":                  name,
                "Email":                 email,
                "Training Provider":     ROLE_BADGE_PROVIDER,
                "Skills Area":           skills_area_clean,
                "Skills Area and Level": skills_area_and_level,
                "Badge Level":           "Role Badge",
                "Date of Award":         date_display,
                "Month":                 month_str,
                "Year":                  year_str,
                "Programme":             programme,
            }
 
            for field, value in new_data.items():
                col_num = sheet_col.get(field)
                if col_num is None:
                    continue
                cell = ws.cell(row=next_row, column=col_num,
                               value=value if value != "" else None)
                if col_num in centre_cols:
                    cell.alignment = Alignment(horizontal="center")
 
            logger.info("    + %s  →  %s  (%s)", email, role_badge_name, date_display)
            _append_log(LOG_FILE, timestamp, email, role_badge_name,
                        "AWARDED", f"Date of Award: {date_display} | {programme}")
 
            # Update in-memory state so the same user isn't awarded twice
            # if they qualify for multiple role badges in the same run.
            # Store BOTH the dict-key form and the bracket form so dedup catches
            # either spelling on the next read.
            users[email.lower()]["existing_role_badges"].add(role_badge_lower)
            users[email.lower()]["existing_role_badges"].add(skills_area_and_level.lower())
 
            next_row     += 1
            new_rows_added += 1
 
        wb.save(path)
        logger.info("  Saved %s", path)
 
    logger.info(
        "Finished. %d role badge row(s) added to '%s'.",
        new_rows_added, MASTERLIST_SHEET,
    )
 
 
# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
 
if __name__ == "__main__":
    _check_and_award(MASTER_FILE)
